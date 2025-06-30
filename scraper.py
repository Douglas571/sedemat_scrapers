import sys
import openpyxl
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
from datetime import datetime, date

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

isDebugging = False

def extraer_numero_comprobante(texto):
    return texto[-5:].replace("°", "").strip()

def extraer_fecha(texto):
    """
    Extrae el día, mes y año de un texto en formato "25 DE ABRIL 2024" y devuelve una fecha.
    
    Args:
    texto (str): La cadena de texto que contiene la fecha.
    
    Returns:
    dict: Un diccionario con las claves 'dia', 'mes', 'anio' y 'fecha'.
    """
    # Diccionario para convertir los nombres de los meses en español a números
    meses = {
        "ENERO": "01", "FEBRERO": "02", "MARZO": "03", "ABRIL": "04",
        "MAYO": "05", "JUNIO": "06", "JULIO": "07", "AGOSTO": "08",
        "SEPTIEMBRE": "09", "OCTUBRE": "10", "NOVIEMBRE": "11", "DICIEMBRE": "12"
    }
    
    # Expresión regular para extraer el día, el mes y el año
    match = re.search(r"(\d{1,2}) DE (\w+) (\d{4})", texto, re.IGNORECASE)
    if match:
        dia = int(match.group(1))
        mes_texto = match.group(2).upper()
        anio = int(match.group(3))
        
        # Convertir el mes en número usando el diccionario
        mes = int(meses.get(mes_texto, "01"))  # Valor predeterminado "01" en caso de no coincidir
        
        try:
            # Crear un objeto de fecha
            fecha = date(anio, mes, dia)
        except ValueError:
            fecha = None  # En caso de fecha inválida

        # Devolver un diccionario con los valores y la fecha
        return fecha
    else:
        return None

def encontrar_monto(hoja):
    """
    Encuentra el valor del monto en la hoja de Excel.
    
    Args:
    hoja (openpyxl.worksheet.worksheet.Worksheet): La hoja de Excel donde buscar el monto.
    
    Returns:
    float or None: El valor del monto si se encuentra, de lo contrario None.
    """
    monto_fila = None
    
    # Iterar sobre las celdas desde A14 hasta A17
    for fila in range(14, 18):  # A14 to A17 (inclusive)
        celda_etiqueta = hoja[f'A{fila}'].value
        if celda_etiqueta and "MONTO:" in celda_etiqueta.upper():  # Normalizar a mayúsculas para comparación
            monto_fila = fila
            break  # Salir del bucle una vez que encontramos la fila

    # Si se encontró la fila del monto, obtener el valor en la columna B de esa fila
    if monto_fila:
        monto_valor = hoja[f'B{monto_fila}'].value
        if monto_valor == 'EXONERADO':
            return 0.0  # Tratamiento especial para el caso 'EXONERADO'
        return float(monto_valor) if monto_valor else None

    return None  # Retornar None si no se encuentra la fila con "MONTO:"

def procesar_excel_y_exportar_excel(archivo_excel, archivo_salida):
    # Cargar el archivo Excel
    libro = openpyxl.load_workbook(archivo_excel, data_only=True)
    
    # Lista para almacenar las liquidaciones
    liquidaciones = []
    solvencias_inmobiliarias = []
    
    # Lista para almacenar los conceptos
    conceptos = []

    # Iterar sobre las hojas del libro, empezando desde la segunda
    for hoja in libro.worksheets[1:]:
        comprobante = {}
        
        # Extraer el número de comprobante
        texto_comprobante = hoja['B8'].value
        # print(hoja.title)
        num_comprobante = extraer_numero_comprobante(texto_comprobante)
        
        # Extraer la fecha
        texto_fecha = hoja['B10'].value
        fecha = extraer_fecha(texto_fecha)
        
        # Extraer otros datos
        razon_social = hoja['C12'].value
        rif_cedula = f"{hoja['H12'].value or ''} {hoja['H13'].value or ''}".strip()
        if hoja['H13'].value != "":
            es_cedula = True
        # en caso de empezar por "PAGO POR: "
        # pago_por = (hoja['C14'].value or '')[12:].strip()

        pago_por = (hoja['A21'].value or '').split("-")[1].strip()
        
        
        # Identify if the settlmeent if for economic licence mantainance 
        description = hoja['C14'].value or ''

        # Check if the description includes "patente", "industria", and "comercio"
        if all(keyword in description.lower() for keyword in ["patente", "industria", "comercio"]):
            if 'mantenimiento' in description.lower():
                pago_por = 'MANTENIMIENTO DE PATENTE DE INDUSTRIA Y COMERCIO'
            elif ('inscripción' in description.lower()) or ('inscripcion' in description.lower()):
                pago_por = 'INSCRIPCION DE PATENTE DE INDUSTRIA Y COMERCIO'

        monto = encontrar_monto(hoja)
        
        # Buscar la fila que contiene "Datos del pago" entre C32 y C36
        datos_del_pago_primera_fila = None
        for fila in range(20, 40):
            valor_celda = hoja[f'A{fila}'].value
            if valor_celda and "DATOS DEL PAGO" in valor_celda:
                datos_del_pago_primera_fila = fila
                break

        if not datos_del_pago_primera_fila:
            print('value not found for: ', hoja.title)

        isExonerated = hoja['B17'].value == 'EXONERADO'


        
        if datos_del_pago_primera_fila:
            # Obtener los datos del pago
            banco = hoja[f'C{datos_del_pago_primera_fila + 1}'].value or ""

            banco = banco.replace("BANCO", "").strip()
            banco = banco.replace("DE", "").strip()
            
            cuenta = hoja[f'C{datos_del_pago_primera_fila + 2}'].value or ""
            cuenta = cuenta.strip()
            
            fecha_pago = hoja[f'C{datos_del_pago_primera_fila + 4}'].value
            referencia = hoja[f'C{datos_del_pago_primera_fila + 5}'].value
            verificado_por = hoja[f'C{datos_del_pago_primera_fila + 6}'].value

            # Crear el registro del comprobante si se encontró el número de comprobante
            if num_comprobante:

                if isDebugging:
                    print(num_comprobante)

                comprobante['num_comprobante'] = num_comprobante
                comprobante['fecha'] = fecha
                comprobante['razon_social'] = razon_social
                comprobante['rif_cedula'] = rif_cedula
                comprobante['pago_por'] = pago_por
                comprobante['monto'] = monto
                comprobante['banco'] = banco

                if (len(cuenta) > 4):
                    comprobante['cuenta'] = cuenta[len(cuenta) - 4:len(cuenta)]
                else:
                    comprobante['cuenta'] = cuenta
                
                comprobante['fecha_pago'] = fecha_pago
                comprobante['referencia'] = referencia
                comprobante['verificado_por'] = verificado_por
                comprobante['es_cedula'] = es_cedula

                if isExonerated: 
                    comprobante['banco'] = 'EXONERADO'
                    comprobante['cuenta'] = 'EXONERADO'
                    comprobante['fecha_pago'] = 'EXONERADO'
                    comprobante['referencia'] = 'EXONERADO'
            
                liquidaciones.append(comprobante)

                conceptos_fila_cabecera = None
                for fila in range(18, 22):
                    valor_celda = hoja[f'A{fila}'].value
                    if valor_celda and "CÓDIGO" in valor_celda:
                        conceptos_fila_cabecera = fila
                        break

                if isDebugging: 
                    print(comprobante)

                # Recolectar conceptos para este num_comprobante
                for fila in range(conceptos_fila_cabecera + 1, 29):  # A21 a A28
                    partida = hoja[f'A{fila}'].value

                    if isDebugging: 
                        print('partida: ', partida)

                    if partida and 'DATOS' in partida: break

                    if partida:
                        monto_concepto = hoja[f'H{fila}'].value
                        concepto = {
                            'partida': partida.split("-")[0].strip(),
                            'descripcion': partida.split("-")[1].strip(),
                            'monto': monto_concepto,
                            'num_comprobante': num_comprobante
                        }
                        conceptos.append(concepto)

        
        # scrap "impuestos sobre la propiedad inmobiliaria"
        
        if 'CATASTRAL' in description:

            parts = []

            if "UBICADA" in description: 
                parts = description.split("UBICADA")

            if "UBICADO" in description: 
                parts = description.split("UBICADO")
            
            if len(parts) > 1:
                address = parts[1].split("ASIGNADA")[0].strip()
                parts = description.split("CATASTRAL")

                if len(parts) > 1:
                    catastral_code = parts[1].split(".")[0].strip()
                    catastral_code = catastral_code.replace('Nº', '').strip()

                    i = len(solvencias_inmobiliarias) + 1

                    operation = ''

                    if "IMPUESTO" in description and "PROPIEDAD" in description and "INMOBILIARIA" in description:
                        operation = 'SOLVENCIA PROPIEDAD INMOBILIARIA'

                    if "ARRENDAMIENTO" in description and "TERRENO" in description:
                        operation = 'ARRENDAMIENTO DE TERRENOS'

                    if "VENTA" in description and "TERRENO" in description:
                        operation = 'VENTA DE TERRENOS'

                    if "ZONIFICACION" in description and "TERRENO" in description:
                        operation = 'ZONIFICACION DE TERRENOS'

                    

                    solvencia = {
                        'n': i,
                        'razon_social': razon_social,
                        'rif_cedula': rif_cedula,
                        'direccion': address,
                        'codigo_catastral': catastral_code,
                        'num_comprobante': num_comprobante,
                        'concepto': operation,
                    }
                    solvencias_inmobiliarias.append(solvencia)

    # Crear un nuevo libro para las liquidaciones y conceptos
    nuevo_libro = openpyxl.Workbook()
    
    # Hoja para las liquidaciones
    nueva_hoja_liquidaciones = nuevo_libro.active
    nueva_hoja_liquidaciones.title = "Liquidaciones"
    
    # Definir los nombres de las columnas para liquidaciones
    campos_liquidaciones = ['razon_social', 'rif_cedula', 'num_comprobante', 'pago_por', 'fecha_pago', 'fecha', 'cuenta', 'banco', 'referencia', 'monto'] #, 'verificado_por', 'es_cedula']
    
    # Escribir la cabecera de liquidaciones
    for col_num, campo in enumerate(campos_liquidaciones, start=1):
        celda = nueva_hoja_liquidaciones.cell(row=1, column=col_num)
        celda.value = campo
    
    # Escribir las filas de liquidaciones
    for fila_num, liquidacion in enumerate(liquidaciones, start=2):
        for col_num, campo in enumerate(campos_liquidaciones, start=1):
            celda = nueva_hoja_liquidaciones.cell(row=fila_num, column=col_num)
            celda.value = liquidacion[campo]
    
    # Crear una tabla para liquidaciones
    max_fila_liquidaciones = nueva_hoja_liquidaciones.max_row
    max_col_liquidaciones = nueva_hoja_liquidaciones.max_column
    rango_tabla_liquidaciones = f"A1:{get_column_letter(max_col_liquidaciones)}{max_fila_liquidaciones}"
    tabla_liquidaciones = Table(displayName="TablaLiquidaciones", ref=rango_tabla_liquidaciones)

    # Estilo de la tabla para liquidaciones
    estilo_liquidaciones = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_liquidaciones.tableStyleInfo = estilo_liquidaciones
    nueva_hoja_liquidaciones.add_table(tabla_liquidaciones)
    
    # Hoja para los conceptos
    nueva_hoja_conceptos = nuevo_libro.create_sheet(title="Conceptos")
    
    # Definir los nombres de las columnas para conceptos
    campos_conceptos = ['partida', 'descripcion', 'monto', 'num_comprobante']
    
    # Escribir la cabecera de conceptos
    for col_num, campo in enumerate(campos_conceptos, start=1):
        celda = nueva_hoja_conceptos.cell(row=1, column=col_num)
        celda.value = campo
    
    # Escribir las filas de conceptos
    for fila_num, concepto in enumerate(conceptos, start=2):
        for col_num, campo in enumerate(campos_conceptos, start=1):
            celda = nueva_hoja_conceptos.cell(row=fila_num, column=col_num)
            celda.value = concepto[campo]
    
    # Crear una tabla para conceptos
    max_fila_conceptos = len(conceptos) + 1  # +1 para contar la fila de cabecera
    max_col_conceptos = nueva_hoja_conceptos.max_column
    rango_tabla_conceptos = f"A1:{get_column_letter(max_col_conceptos)}{max_fila_conceptos}"
    tabla_conceptos = Table(displayName="TablaConceptos", ref=rango_tabla_conceptos)

    # Estilo de la tabla para conceptos
    estilo_conceptos = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_conceptos.tableStyleInfo = estilo_conceptos
    nueva_hoja_conceptos.add_table(tabla_conceptos)


    #### Hoja para las solvencias inmobiliarias ###########
    nueva_hoja_solvencias_inmobiliarias = nuevo_libro.create_sheet(title="SolvenciasInmobiliarias")
    
    # Definir los nombres de las columnas para conceptos
    campos_solvencias_inmobiliarias = ['n', 'razon_social', 'rif_cedula', 'codigo_catastral', 'direccion', 'num_comprobante', 'concepto']
    
    # Escribir la cabecera de solvencias inmobiliarias
    for col_num, campo in enumerate(campos_solvencias_inmobiliarias, start=1):
        celda = nueva_hoja_solvencias_inmobiliarias.cell(row=1, column=col_num)
        celda.value = campo
    
    # Escribir las filas de solvencias inmobiliarias
    for fila_num, solvencia in enumerate(solvencias_inmobiliarias, start=2):
        for col_num, campo in enumerate(campos_solvencias_inmobiliarias, start=1):
            celda = nueva_hoja_solvencias_inmobiliarias.cell(row=fila_num, column=col_num)
            celda.value = solvencia[campo]

    
    # Guardar el archivo Excel
    nuevo_libro.save(archivo_salida)

# Ejemplo de uso
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python3 scraper.py <input_file> <output_file.xlsx>")
        sys.exit(1)

    archivo_entrada = sys.argv[1]
    archivo_salida = sys.argv[2]
    procesar_excel_y_exportar_excel(archivo_entrada, archivo_salida)
