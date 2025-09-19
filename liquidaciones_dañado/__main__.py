"""
i need an script to iterate over excel sheets and extract the data i need

the output table will have the following columns 
razon_social =>  C12
rif_cedula => H12
num_comprobante => B8
pago_por => null for now
fecha_pago =>
fecha => B10
cuenta =>
banco =>
referencia =>
monto => B17

create a list of dictionaries

for each sheet

  create a dictionary 

  look for 
    razon_social =>  C12
    rif_cedula => H12
    num_comprobante => B8
    fecha => B10
    monto => B17

  iterate over A column until you find "DATOS DEL PAGO", then asigne the index to paymentDataStartIndex

    fecha_pago => C,paymentDataStartIndex+4
    cuenta => C,paymentDataStartIndex+2
    banco => C,paymentDataStartIndex+1
    referencia => C,paymentDataStartIndex+5

  push the dictionary 

  
print the list of dictionaries

"""



from datetime import date
import re
import sys
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def extract_data_from_excel(file_path):
    """
    Extract data from all sheets in an Excel file according to specified cell locations
    """
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)
    
    data_list = []
    
    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Create dictionary for this sheet
        sheet_data = {}
        
        try:
            # Extract fixed cell values
            sheet_data['razon_social'] = ws['C12'].value
            
            cell_value = ws['H12'].value
            if not cell_value:
                sheet_data['rif_cedula'] = ws['H13'].value
            else:
                sheet_data['rif_cedula'] = cell_value

            cell_value = ws['B8'].value
            if cell_value:
                match = re.search(r"COMPROBANTE DE INGRESO N°(\d+)", str(cell_value), re.IGNORECASE)
                if match:
                    sheet_data['num_comprobante'] = match.group(1)

            cell_value = ws['B10'].value
            if cell_value:
                match = re.search(r"(\d{1,2}) DE (\w+) (\d{4})", str(cell_value), re.IGNORECASE)
                if match:
                    dia = int(match.group(1))
                    mes_nombre = match.group(2)
                    anio = int(match.group(3))
                    meses = {
                        'ENERO': 1,
                        'FEBRERO': 2,
                        'MARZO': 3,
                        'ABRIL': 4,
                        'MAYO': 5,
                        'JUNIO': 6,
                        'JULIO': 7,
                        'AGOSTO': 8,
                        'SEPTIEMBRE': 9,
                        'OCTUBRE': 10,
                        'NOVIEMBRE': 11,
                        'DICIEMBRE': 12
                    }
                    mes = meses[mes_nombre]
                    sheet_data['fecha'] = date(anio, mes, dia)
                    
            sheet_data['monto'] = ws['B17'].value
            
            # Initialize payment data fields
            sheet_data['pago_por'] = None
            sheet_data['fecha_pago'] = None
            sheet_data['cuenta'] = None
            sheet_data['banco'] = None
            sheet_data['referencia'] = None
            
            # Find "DATOS DEL PAGO" in column A
            paymentDataStartIndex = None
            for row in range(1, 100):  # Search first 100 rows
                cell_value = ws[f'A{row}'].value
                if cell_value and "DATOS DEL PAGO" in str(cell_value).upper():
                    paymentDataStartIndex = row
                    break
            
            # Extract payment data if found
            if paymentDataStartIndex:
                try:
                    sheet_data['fecha_pago'] = ws[f'C{paymentDataStartIndex + 4}'].value
                    sheet_data['cuenta'] = ws[f'C{paymentDataStartIndex + 2}'].value
                    sheet_data['banco'] = ws[f'C{paymentDataStartIndex + 1}'].value
                    sheet_data['referencia'] = ws[f'C{paymentDataStartIndex + 5}'].value
                except:
                    print(f"Warning: Could not extract payment data from sheet '{sheet_name}'")
            
            # Add sheet name for reference
            sheet_data['sheet_name'] = sheet_name
            
            data_list.append(sheet_data)
            
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {e}")
            continue
    
    return data_list

def save_to_excel(data_list, output_file):
    """
    Save the list of dictionaries to an Excel file
    """
    # Convert to DataFrame
    df = pd.DataFrame(data_list)
    
    # Reorder columns as specified
    columns_order = [
        'razon_social', 'rif_cedula', 'num_comprobante', 'pago_por', 
        'fecha_pago', 'fecha', 'cuenta', 'banco', 'referencia', 'monto', 'sheet_name'
    ]
    
    # Only include columns that exist in the data
    existing_columns = [col for col in columns_order if col in df.columns]
    
    # Save to Excel
    df[existing_columns].to_excel(output_file, index=False)
    print(f"Data saved to {output_file}")

def main():
    # Get input file from command line argument
    if len(sys.argv) < 2:
        print("Error: You must provide the path to your Excel file as a command line argument.")
        return
    
    input_file = sys.argv[1]
    
    # Verify file exists
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found.")
        return
    
    # Extract data
    print("Extracting data from sheets...")
    extracted_data = extract_data_from_excel(input_file)
    
    if not extracted_data:
        print("No data was extracted. Please check the file format.")
        return
    
    # Output file path
    output_file = os.path.splitext(input_file)[0] + "_extracted.xlsx"
    
    # Save results
    save_to_excel(extracted_data, output_file)


if __name__ == "__main__":
    main()