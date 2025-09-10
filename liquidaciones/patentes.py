import openpyxl
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
from datetime import datetime, date

import locale
locale.setlocale( locale.LC_ALL, '' )

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def extraer_numero_comprobante(hoja):
    codigo = hoja['E8'].value
    
    if codigo: 
        codigo = codigo.strip().replace("Nº: ", "")

    return codigo

def extraer_fecha(hoja):
    fecha = hoja['B9'].value

    if fecha:
        fecha = fecha.replace("PUERTO CUMAREBO ", "")

    return fecha



def encontrar_monto(hoja):
    """
    Encuentra el valor del monto en la hoja de Excel.
    
    Args:
    hoja (openpyxl.worksheet.worksheet.Worksheet): La hoja de Excel donde buscar el monto.
    
    Returns:
    float or None: El valor del monto si se encuentra, de lo contrario None.
    """
    monto_fila = None
    
    # Iterar sobre las celdas desde A14 hasta A17
    for fila in range(14, 18):  # A14 to A17 (inclusive)
        celda_etiqueta = hoja[f'A{fila}'].value
        if celda_etiqueta and "MONTO:" in celda_etiqueta.upper():  # Normalizar a mayúsculas para comparación
            monto_fila = fila
            break  # Salir del bucle una vez que encontramos la fila

    # Si se encontró la fila del monto, obtener el valor en la columna B de esa fila
    if monto_fila:
        monto_valor = hoja[f'B{monto_fila}'].value
        if monto_valor == 'EXONERADO':
            return 0.0  # Tratamiento especial para el caso 'EXONERADO'
        return float(monto_valor) if monto_valor else None

    return None  # Retornar None si no se encuentra la fila con "MONTO:"

def extraer_monto(hoja):
    monto = hoja['F19'].value

    print(monto)

    if monto and type(monto) is str:
        monto = monto.strip()
        monto = monto.replace(" BS", "")
        monto = monto.replace("BS", "")
        monto = locale.delocalize(monto)

        if (monto):
            monto = float(monto)

    return monto

def procesar_excel_y_exportar_excel(archivo_excel, archivo_salida):
    # Cargar el archivo Excel
    libro = openpyxl.load_workbook(archivo_excel, data_only=True)
    
    patentes = []

    for hoja in libro.worksheets:
        patente = {}

        codigo = extraer_numero_comprobante(hoja)

        cedula = hoja['F14'].value
        razon_social = hoja['C14'].value

        placa = hoja['C17'].value

        monto = extraer_monto(hoja)
        
        patente["codigo"] = codigo
        patente["cedula"] = cedula
        patente["razon_social"] = razon_social
        patente["placa"] = placa
        patente["monto"] = monto

        patente["fecha"] = extraer_fecha(hoja)
        patente["referencia"] = ""

        patente["marca"] = hoja['F15'].value
        patente["modelo"] = hoja['C15'].value
        patente["año"] = hoja['C16'].value
        patente["color"] = hoja['F16'].value
        patente["uso"] = hoja['F17'].value

        # # print(codigo, " ", cedula, " ", razon_social, " ", placa, " ", monto)
        # print(patente['fecha'])

        patentes.append(patente)

    nuevo_libro = openpyxl.Workbook()
    
    # Hoja para las liquidaciones
    nueva_hoja_patentes = nuevo_libro.active
    nueva_hoja_patentes.title = "patentes"
    
    # Definir los nombres de las columnas para liquidaciones
    campos_patente = ['codigo', 'razon_social', 'cedula', 'placa', 'fecha', 'monto', 'referencia', 'marca', 'modelo', 'año', 'color', 'uso'] #, 'verificado_por', 'es_cedula']
    
    # Escribir la cabecera de liquidaciones
    for col_num, campo in enumerate(campos_patente, start=1):
        celda = nueva_hoja_patentes.cell(row=1, column=col_num)
        celda.value = campo
    
    # Escribir las filas de liquidaciones
    for fila_num, patente in enumerate(patentes, start=2):
        for col_num, campo in enumerate(campos_patente, start=1):
            celda = nueva_hoja_patentes.cell(row=fila_num, column=col_num)
            celda.value = patente[campo]

    nuevo_libro.save(archivo_salida)

import sys

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Error: se necesitan 2 argumentos")
        sys.exit(1)

    archivo_entrada = sys.argv[1]
    archivo_salida = sys.argv[2]

    procesar_excel_y_exportar_excel(archivo_entrada, archivo_salida)
