import time
import locale
import pandas as pd
import os
from datetime import datetime
import openpyxl

locale.setlocale(locale.LC_ALL, 'es_VE.utf8')

# this program will check taht that the payments in account statement files, are in the settlement files for that month 
# 
# get for console input 
#   month (it can be a number from 1 to 12) 
#   year (it can be a number from 2020 to current year)

startTime = time.time()

# ask user for month and year
month = input("Enter month (1-12): ")
year = input("Enter year (2020-current): ")



# validate month and year
current_year = datetime.now().year
current_month = datetime.now().month

if not year:
    year = current_year
else:
    year = int(year)

if not month:
    month = current_month
else:
    month = int(month)


if not (1 <= month <= 12):
    raise ValueError("Month must be between 1 and 12")

if not (2020 <= year <= current_year):
    raise ValueError(f"Year must be between 2020 and {current_year}")

# format month and year for filenames
mm = f"{month:02d}"
yy = str(year)[-2:]

# --------------------------
#          LOAD FASE
# --------------------------

# after you get the month and year, load the account statements for the accounts 1892 and 9290
account_statements_dir = "./datos/account_statements"

file_9290 = os.path.join(account_statements_dir, f"{mm}-{yy}-9290.xlsx")
file_1892 = os.path.join(account_statements_dir, f"{mm}-{yy}-1892.xlsx")

# for 9290
#  look into the sheet "Table 2"
#  it has the following columns
#  1. Fecha
#  2. Referencia
#  3. Código 
#  4. Descripción
#  5. Débito
#  6. Crédito
#  7. Saldo
df_9290 = pd.read_excel(file_9290, sheet_name="Table 2")

# for 1892
#  look into the sheet "data"
#  it has the following columns
#  1. fecha
#  2. referencia
#  3. concepto 
#  4. saldo
#  5. month
#  6. tipoMovimiento
#  7. rif
#  8. numeroCuenta
df_1892 = pd.read_excel(file_1892, sheet_name="data")

# for biopago transactions, load the file from ./datos/account_statements/{MM}-{YY}-biopago.xlsx
# it has the following columns
# Nro.	Fecha	Instrumento	Emisor	Monto	Equipo	Lote	Cédula Pagador	Resultado	Autorización
biopago_file = f"./datos/account_statements/{mm}-{yy}-biopago.xlsx"
df_biopago = pd.read_excel(biopago_file, skiprows=[0], header=None, names=[
    "number",
    "date",
    "instrument",
    "issuer",
    "amount",
    "equipment",
    "lot",
    "payer_id",
    "result",
    "authorization"
])

# print("Biopago transactions loaded:", df_biopago.shape)
# print(df_biopago.to_string())

# load the settlments 
# the file is in ./datos/settlements/cuadro-{MM}-{YY}.xlsx
# it has the following columns
# 1. razon_social
# 2. rif_cedula
# 3. num_comprobante
# 4. pago_por
# 5. fecha_pago
# 6. fecha
# 7. cuenta
# 8. banco
# 9. referncia
# 10. monto
# settlements_file = f"./datos/settlements/cuadro-{mm}-{yy}.xlsx"
settlements_file = f"./datos/settlements/cuadro_to_use.xlsx"
df_settlements = pd.read_excel(settlements_file)

# in this case, remove all the settlements that has "EXONERADO" in refernece column
# df_settlements = df_settlements[~df_settlements['referncia'].astype(str).str.contains("EXONERADO", case=False, na=False)]

# now you have:
# df_9290 → account 9290 statement
# df_1892 → account 1892 statement
# df_settlements → settlements (filtered)

# here you would implement the logic to check that payments in account statements are in settlements
# (matching by reference, amount, or other criteria depending on your business rules)

print ('---- load data / %s seconds ----' % (time.time() - startTime))

print("Account 9290 statement loaded:", df_9290.shape)
print("Account 1892 statement loaded:", df_1892.shape)
print("Settlements loaded (filtered):", df_settlements.shape)

# print("Data in account 1892 statement:")
# print(df_1892.to_string())

# print("Data in account 9290 statement:")
# print(df_9290.to_string())

# ------------------------------------------
#             NORMALIZATION FASE 
# ------------------------------------------

# the common structure for data will be
#  1. reference 
#  2. amount
#  3. date
#  4. bank
#  5. account_number 

# map the data in df_9290 and df_1892 into the common structure
# 9290 => BDT
# 1892 => BANCO DE VENEZUELA

normalization_start_time = time.time()

# normalize df_9290
df_9290_norm = pd.DataFrame({
    "reference": df_9290["Referencia"],
    "amount": df_9290["Crédito"].fillna(0) - df_9290["Débito"].fillna(0),  # positive = credit, negative = debit
    "date": pd.to_datetime(df_9290["Fecha"], format="%d-%m-%Y", errors="coerce"),
    "bank": "BDT",
    "account_number": "9290",
    "description": df_9290["Descripción"],
    "settlementCode": '',
    "settlementDate": None
})

# normalize df_1892
df_1892_norm = pd.DataFrame({
    "reference": df_1892["referencia"],
    "amount": df_1892["monto"],  # assuming saldo is the transaction amount
    "date": pd.to_datetime(df_1892["fecha"], format="%d/%m/%Y", errors="coerce"),
    "bank": "BANCO DE VENEZUELA",
    "account_number": "1892",
    "description": df_1892["concepto"],
    "settlementCode": '',
    "settlementDate": None
})

df_biopago_norm = pd.DataFrame({
    "reference": df_biopago["number"],
    "amount": df_biopago["amount"],
    "date": pd.to_datetime(df_biopago["date"], format="%d/%m/%Y", errors="coerce"),
    "bank": "BIOPAGO",
    "account_number": "1892",
    "description": df_biopago["equipment"],
    "settlementCode": '',
    "settlementDate": None
})

# merge then in a single object
payments = pd.concat([df_9290_norm, df_1892_norm, df_biopago_norm], ignore_index=True)

# print the list of payments
# print("Normalized Payments:")
# print(payments.to_string())


print ('---- normalized data / %s seconds ----' % (time.time() - normalization_start_time))

# ------------------------------------------
#             CONSOLIDATION FASE 
# ------------------------------------------

consolidation_start_time = time.time()

not_settled_payments = []

paymentsDict = payments.to_dict(orient="records")
filteredPaymentsDict = []

# for each payment
for index, payment in payments.iterrows():

  # for each settlement
  found = False
  amount = 0

  if isinstance(payment["amount"], str):
    amount = locale.atof(payment["amount"])
    paymentsDict[index]["amount"] = amount

  description = str(payment["description"]).lower()

#   if not any(word in description for word in ["saldo inicial", "mantenimiento", "comision", "emision", "cargo", 'servicio'  ]) and amount > 0:
#     continue

  payment_reference = str(payment["reference"]).strip().split(".")[0]


  for index_settlement, settlement in df_settlements.iterrows():
    # if payment reference is included in settlement reference, continue with another payment
    settlement_references = str(settlement["referencia"]).strip().split("-")

    for st in settlement_references:
    
      string_to_check = st

    #   if "CUMAREBO" in str(payment["description"]):
    #     string_to_check = st[-4:]

      if payment_reference.endswith(string_to_check) and len(string_to_check) > 0:

        if "CUMAREBO" in str(payment["description"]):
            print(string_to_check, ' ', payment_reference, ' / ', str(int(settlement["num_comprobante"])))

        found = True
        paymentsDict[index]["settlementCode"] = str(int(settlement["num_comprobante"]))
        paymentsDict[index]["settlementDate"] = datetime.strptime(str(settlement["fecha"]), "%Y-%m-%d %H:%M:%S").date()

        filteredPaymentsDict.append(paymentsDict[index])
        


# # ------------------------------------------
# #        BIOPAGO CONSOLIDATION FASE
# # ------------------------------------------


# # get a list of payments with "BIOPAGO" in description
# biopago_payments = payments[payments["description"].str.contains("BIOPAGO", case=False)]
# # get a list of payments with "LIQUIDACION TDD BIOPAGOBDV" in description
# liquidation_payments = payments[payments["description"].str.contains("LIQUIDACION TDD BIOPAGOBDV", case=False)]

# # 



print ('---- consolidated data / %s seconds ----' % (time.time() - consolidation_start_time))

# ------------------------------------------
#                STORE FASE 
# ------------------------------------------

store_start_time = time.time()

# filter from paymentDict all the payments that contains the following words in description
newPaymentsDict = []

for payment in paymentsDict: 
   description = str(payment["description"]).lower()

   if not any(word in description for word in ["saldo inicial", "mantenimiento", "comision", "emision", "cargo", 'servicio'  ]) and payment["amount"] > 0:
      newPaymentsDict.append(payment)

toPrintData = pd.DataFrame(newPaymentsDict)
toPrintData.columns = [
    "referencia",
    "monto",
    "fecha",
    "banco",
    "numero_cuenta",
    "descripcion",
    "codigo_liquidacion",
    "fecha_liquidacion"
]
# print(toPrintData.to_string())

toPrintData.to_excel(f"./datos/payments_{mm}_{yy}.xlsx", index=False)
# print(f"File ./datos/payments_{mm}_{yy}.xlsx generated with {len(payments)} payments")

print ('---- store data / %s seconds ----' % (time.time() - store_start_time))


# generate an excel file with the payments not settled
# not_settled_payments_df = pd.DataFrame(not_settled_payments)
# not_settled_payments_file = f"./datos/not_settled_payments_{mm}_{yy}.xlsx"
# not_settled_payments_df.to_excel(not_settled_payments_file, index=False)
# print(f"File {not_settled_payments_file} generated with {len(not_settled_payments)} payments not settled")

# open the file
wb = openpyxl.load_workbook(filename=f"./datos/payments_{mm}_{yy}.xlsx")
ws = wb.active

# iterate over each row
for row in ws.iter_rows(values_only=False):
    # if row G column is not empty, fill the row with red color, otherwise, fill green
    if not row[6].value:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color='FFC7CE', fill_type='solid')
    else:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color='C6EFCE', fill_type='solid')

    # if the row D column has "BIOPAGO" and G column is not empty, fill with yellow colors
    if row[3].value == "BIOPAGO" and row[6].value:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color='FFEB9C', fill_type='solid')

# save the file
wb.save(filename=f"./datos/payments_{mm}_{yy}.xlsx")
